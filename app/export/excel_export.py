"""Excel export for saved email requests."""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


EXPORT_DIR = Path("exports")
HEADERS = (
    "Email",
    "Компания",
    "Дата",
    "Тема",
    "Категория",
    "Приоритет",
    "Бюджет",
    "Срок",
)


def _display_value(value: Any, fallback: str = "") -> Any:
    return value if value not in (None, "") else fallback


def export_requests_to_excel(records: list[dict]) -> str:
    """Create a formatted Excel workbook with the request history."""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"requests_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    file_path = EXPORT_DIR / filename

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Обращения"
    worksheet.append(HEADERS)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for record in records:
        worksheet.append(
            (
                _display_value(record.get("email"), "не указан"),
                _display_value(record.get("company"), "не указана"),
                _display_value(record.get("created_at")),
                _display_value(record.get("subject")),
                _display_value(record.get("category")),
                _display_value(record.get("priority")),
                _display_value(record.get("budget")),
                _display_value(record.get("deadline")),
            )
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    workbook.save(file_path)
    return str(file_path)

